#!/usr/bin/env python3
"""
Test script for File API.

Tests:
1. File storage module
2. File upload/download
3. Session management
4. File processing with Cape

Usage:
    python test_file_api.py
"""

import asyncio
import sys
import tempfile
from pathlib import Path

# Add project to path
sys.path.insert(0, str(Path(__file__).parent))


async def test_file_storage():
    """Test FileStorage module."""
    print("\n=== Testing FileStorage ===")

    from api.storage import FileStorage, StorageConfig, FileStatus

    # Create storage with temp directory
    with tempfile.TemporaryDirectory() as temp_dir:
        config = StorageConfig(
            base_dir=Path(temp_dir),
            max_file_size_mb=10,
            retention_hours=1,
        )
        storage = FileStorage(config)
        await storage.initialize()

        try:
            # Test 1: Upload file
            print("  Test 1: Upload file...", end=" ")
            test_content = b"Hello, World! This is a test file."
            metadata = await storage.upload(
                content=test_content,
                filename="test.txt",
                session_id="session-1",
            )
            assert metadata.file_id
            assert metadata.original_name == "test.txt"
            assert metadata.size_bytes == len(test_content)
            assert metadata.status == FileStatus.UPLOADED
            print("✓")

            # Test 2: Download file
            print("  Test 2: Download file...", end=" ")
            content, meta = await storage.download(metadata.file_id)
            assert content == test_content
            assert meta.file_id == metadata.file_id
            print("✓")

            # Test 3: Get metadata
            print("  Test 3: Get metadata...", end=" ")
            meta = await storage.get_metadata(metadata.file_id)
            assert meta is not None
            assert meta.original_name == "test.txt"
            print("✓")

            # Test 4: Update status
            print("  Test 4: Update status...", end=" ")
            updated = await storage.update_status(
                metadata.file_id,
                FileStatus.PROCESSING,
                cape_id="test-cape"
            )
            assert updated.status == FileStatus.PROCESSING
            assert updated.cape_id == "test-cape"
            print("✓")

            # Test 5: Upload multiple files
            print("  Test 5: Multiple files...", end=" ")
            await storage.upload(b"File 2", "file2.txt", "session-1")
            await storage.upload(b"File 3", "file3.txt", "session-1")
            files = await storage.list_session_files("session-1")
            assert len(files) == 3
            print("✓")

            # Test 6: Save output file
            print("  Test 6: Save output...", end=" ")
            output = await storage.save_output(
                content=b"Output content",
                filename="output.txt",
                session_id="session-1",
                source_file_id=metadata.file_id,
                cape_id="test-cape",
            )
            assert output.is_output
            assert output.source_file_id == metadata.file_id
            print("✓")

            # Test 7: Delete file
            print("  Test 7: Delete file...", end=" ")
            success = await storage.delete_file(metadata.file_id)
            assert success
            meta = await storage.get_metadata(metadata.file_id)
            assert meta is None
            print("✓")

            # Test 8: Storage stats
            print("  Test 8: Storage stats...", end=" ")
            stats = await storage.get_stats()
            assert stats["total_files"] >= 2
            print("✓")

            # Test 9: Delete session
            print("  Test 9: Delete session...", end=" ")
            deleted = await storage.delete_session("session-1")
            assert deleted >= 2
            files = await storage.list_session_files("session-1")
            assert len(files) == 0
            print("✓")

            print("  FileStorage: All tests passed!")

        finally:
            await storage.shutdown()


async def test_file_validation():
    """Test file validation."""
    print("\n=== Testing File Validation ===")

    from api.storage import (
        FileStorage,
        StorageConfig,
        FileTooLargeError,
        InvalidFileTypeError,
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        config = StorageConfig(
            base_dir=Path(temp_dir),
            max_file_size_mb=1,  # 1MB limit
            allowed_extensions=[".txt", ".pdf"],
        )
        storage = FileStorage(config)
        await storage.initialize()

        try:
            # Test 1: Invalid file type
            print("  Test 1: Invalid file type...", end=" ")
            try:
                await storage.upload(b"content", "file.exe")
                assert False, "Should have raised InvalidFileTypeError"
            except InvalidFileTypeError:
                pass
            print("✓")

            # Test 2: File too large
            print("  Test 2: File too large...", end=" ")
            large_content = b"x" * (2 * 1024 * 1024)  # 2MB
            try:
                await storage.upload(large_content, "large.txt")
                assert False, "Should have raised FileTooLargeError"
            except FileTooLargeError:
                pass
            print("✓")

            # Test 3: Valid file
            print("  Test 3: Valid file...", end=" ")
            metadata = await storage.upload(b"Valid content", "valid.txt")
            assert metadata is not None
            print("✓")

            print("  File Validation: All tests passed!")

        finally:
            await storage.shutdown()


async def test_api_schemas():
    """Test API schemas."""
    print("\n=== Testing API Schemas ===")

    from datetime import datetime, timedelta
    from api.storage import FileMetadata, FileStatus
    from api.routes.files import FileResponse

    # Test FileResponse.from_metadata
    print("  Test 1: FileResponse schema...", end=" ")

    now = datetime.utcnow()
    metadata = FileMetadata(
        file_id="test-id",
        original_name="test.xlsx",
        stored_name="test-id.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=1024,
        checksum="abc123",
        status=FileStatus.UPLOADED,
        session_id="session-1",
        created_at=now,
        expires_at=now + timedelta(hours=24),
    )

    response = FileResponse.from_metadata(metadata)
    assert response.file_id == "test-id"
    assert response.original_name == "test.xlsx"
    assert response.status == "uploaded"
    assert "/api/files/test-id" in response.download_url
    print("✓")

    print("  API Schemas: All tests passed!")


async def test_excel_file():
    """Test Excel file handling."""
    print("\n=== Testing Excel File ===")

    from api.storage import FileStorage, StorageConfig

    with tempfile.TemporaryDirectory() as temp_dir:
        config = StorageConfig(base_dir=Path(temp_dir))
        storage = FileStorage(config)
        await storage.initialize()

        try:
            # Create a simple Excel file
            print("  Test 1: Create Excel file...", end=" ")
            try:
                from openpyxl import Workbook
                import io

                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'Name'
                ws['B1'] = 'Value'
                ws['A2'] = 'Test'
                ws['B2'] = 42

                buffer = io.BytesIO()
                wb.save(buffer)
                xlsx_content = buffer.getvalue()

                print("✓")

                # Upload Excel file
                print("  Test 2: Upload Excel...", end=" ")
                metadata = await storage.upload(
                    content=xlsx_content,
                    filename="test.xlsx",
                    session_id="excel-session",
                )
                assert metadata.content_type in [
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/octet-stream",
                ]
                print("✓")

                # Download and verify
                print("  Test 3: Download Excel...", end=" ")
                content, _ = await storage.download(metadata.file_id)
                assert len(content) > 0

                # Verify content
                from openpyxl import load_workbook
                buffer = io.BytesIO(content)
                wb = load_workbook(buffer)
                ws = wb.active
                assert ws['A1'].value == 'Name'
                assert ws['B2'].value == 42
                print("✓")

                print("  Excel File: All tests passed!")

            except ImportError:
                print("(skipped - openpyxl not installed)")

        finally:
            await storage.shutdown()


async def main():
    """Run all tests."""
    print("=" * 60)
    print("File API Tests")
    print("=" * 60)

    try:
        await test_file_storage()
        await test_file_validation()
        await test_api_schemas()
        await test_excel_file()

        print("\n" + "=" * 60)
        print("ALL FILE API TESTS PASSED!")
        print("=" * 60)

    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
