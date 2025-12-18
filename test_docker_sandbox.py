#!/usr/bin/env python3
"""
Test script for Docker Sandbox.

Tests:
1. Docker availability check
2. Base image build
3. Container lifecycle management
4. Code execution
5. File I/O
6. Timeout handling
7. Resource monitoring

Usage:
    python test_docker_sandbox.py [--build-image]
"""

import asyncio
import sys
import time
from pathlib import Path

# Add project to path
sys.path.insert(0, str(Path(__file__).parent))


async def check_docker():
    """Check if Docker is available."""
    print("\n=== Checking Docker Availability ===")

    try:
        from cape.runtime.sandbox.docker_sandbox import check_docker_available
        available, message = await check_docker_available()

        if available:
            print(f"  ✓ {message}")
            return True
        else:
            print(f"  ✗ {message}")
            return False
    except ImportError as e:
        print(f"  ✗ Import error: {e}")
        return False


async def build_image():
    """Build the base Docker image."""
    print("\n=== Building Base Image ===")

    from cape.runtime.sandbox.docker_sandbox import build_base_image

    print("  Building image (this may take a few minutes)...")
    start = time.time()
    success = await build_base_image(force=False)
    elapsed = time.time() - start

    if success:
        print(f"  ✓ Image ready ({elapsed:.1f}s)")
    else:
        print(f"  ✗ Failed to build image")

    return success


async def test_basic_execution():
    """Test basic code execution in Docker."""
    print("\n=== Testing Basic Execution ===")

    from cape.runtime.sandbox import (
        DockerSandbox,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(
        type=SandboxType.DOCKER,
        timeout_seconds=30,
        max_memory_mb=256,
        max_cpu_percent=50,
    )
    sandbox = DockerSandbox(config)
    await sandbox.setup()

    try:
        # Test 1: Simple calculation
        print("  Test 1: Simple calculation...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="result = 2 + 2"
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == 4, f"Expected 4, got {response.output}"
        print("✓")

        # Test 2: With arguments
        print("  Test 2: With arguments...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="result = args['x'] * args['y']",
            args={"x": 7, "y": 8},
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == 56, f"Expected 56, got {response.output}"
        print("✓")

        # Test 3: Using imports
        print("  Test 3: Using imports...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
import json
data = {"name": "test", "value": 42}
result = json.dumps(data)
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert "test" in response.output
        print("✓")

        # Test 4: Stdout capture
        print("  Test 4: Stdout capture...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
print("Hello from Docker!")
result = "done"
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert "Hello from Docker!" in response.stdout
        print("✓")

        # Test 5: Error handling
        print("  Test 5: Error handling...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="result = 1 / 0"
        ))
        assert not response.success, "Should have failed"
        assert "ZeroDivisionError" in response.stderr or "division" in str(response.error).lower()
        print("✓")

        print("  Basic Execution: All tests passed!")

    finally:
        await sandbox.cleanup()


async def test_file_io():
    """Test file input/output in Docker."""
    print("\n=== Testing File I/O ===")

    from cape.runtime.sandbox import (
        DockerSandbox,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(type=SandboxType.DOCKER)
    sandbox = DockerSandbox(config)
    await sandbox.setup()

    try:
        # Test 1: File input
        print("  Test 1: File input...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
content = Path("input.txt").read_text()
result = len(content)
""",
            files={"input.txt": b"Hello World!"},
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == 12, f"Expected 12, got {response.output}"
        print("✓")

        # Test 2: File output
        print("  Test 2: File output...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
Path("output.txt").write_text("Generated content")
result = "done"
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert "output.txt" in response.files_created
        assert response.files_created["output.txt"] == b"Generated content"
        print("✓")

        # Test 3: Multiple files
        print("  Test 3: Multiple files...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
import json

# Read input
data1 = Path("data1.json").read_text()
data2 = Path("data2.json").read_text()

# Process
combined = {"merged": [json.loads(data1), json.loads(data2)]}

# Write output
Path("result.json").write_text(json.dumps(combined))
result = len(combined["merged"])
""",
            files={
                "data1.json": b'{"a": 1}',
                "data2.json": b'{"b": 2}',
            },
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == 2
        assert "result.json" in response.files_created
        print("✓")

        # Test 4: Subdirectory output
        print("  Test 4: Subdirectory output...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
out_dir = Path("output")
out_dir.mkdir(exist_ok=True)
(out_dir / "file1.txt").write_text("content1")
(out_dir / "file2.txt").write_text("content2")
result = "done"
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert "output/file1.txt" in response.files_created or "output\\file1.txt" in response.files_created
        print("✓")

        print("  File I/O: All tests passed!")

    finally:
        await sandbox.cleanup()


async def test_timeout():
    """Test timeout handling."""
    print("\n=== Testing Timeout ===")

    from cape.runtime.sandbox import (
        DockerSandbox,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(
        type=SandboxType.DOCKER,
        timeout_seconds=3,  # Short timeout
    )
    sandbox = DockerSandbox(config)
    await sandbox.setup()

    try:
        print("  Test: Timeout handling (may take ~3s)...", end=" ")
        start = time.time()

        response = await sandbox.execute(ExecutionRequest(
            code="""
import time
time.sleep(10)
result = "Should not reach"
"""
        ))

        elapsed = time.time() - start

        assert not response.success, "Should have timed out"
        assert "timeout" in response.error.lower()
        assert elapsed < 6, f"Took too long: {elapsed:.1f}s"
        print(f"✓ (timed out in {elapsed:.1f}s)")

        # Verify container is still usable after timeout
        print("  Test: Container recovery...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="result = 'recovered'"
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == "recovered"
        print("✓")

        print("  Timeout: All tests passed!")

    finally:
        await sandbox.cleanup()


async def test_package_installation():
    """Test package installation in Docker."""
    print("\n=== Testing Package Installation ===")

    from cape.runtime.sandbox import (
        DockerSandbox,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(
        type=SandboxType.DOCKER,
        network_enabled=True,  # Need network for pip
    )
    sandbox = DockerSandbox(config)
    await sandbox.setup()

    try:
        # Test using pre-installed package (openpyxl)
        print("  Test 1: Using pre-installed package...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
result = ws['A1'].value
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == "Hello"
        print("✓")

        # Test pandas (also pre-installed)
        print("  Test 2: Using pandas...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="""
import pandas as pd
df = pd.DataFrame({'a': [1, 2, 3], 'b': [4, 5, 6]})
result = df['a'].sum()
"""
        ))
        assert response.success, f"Failed: {response.error}"
        assert response.output == 6
        print("✓")

        print("  Package Installation: All tests passed!")

    finally:
        await sandbox.cleanup()


async def test_resource_stats():
    """Test resource statistics."""
    print("\n=== Testing Resource Stats ===")

    from cape.runtime.sandbox import (
        DockerSandbox,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(
        type=SandboxType.DOCKER,
        max_memory_mb=256,
    )
    sandbox = DockerSandbox(config)
    await sandbox.setup()

    try:
        # Run some computation
        await sandbox.execute(ExecutionRequest(
            code="""
import numpy as np
arr = np.random.rand(1000, 1000)
result = arr.sum()
"""
        ))

        # Get stats
        print("  Test: Get container stats...", end=" ")
        stats = await sandbox.get_container_stats()

        if stats:
            print(f"✓")
            print(f"    CPU: {stats['cpu_percent']}%")
            print(f"    Memory: {stats['memory_usage_mb']:.1f}MB / {stats['memory_limit_mb']:.1f}MB ({stats['memory_percent']:.1f}%)")
        else:
            print("✗ (stats not available)")

        print("  Resource Stats: Test completed!")

    finally:
        await sandbox.cleanup()


async def test_sandbox_manager():
    """Test SandboxManager with Docker."""
    print("\n=== Testing SandboxManager with Docker ===")

    from cape.runtime.sandbox import (
        SandboxManager,
        SandboxConfig,
        SandboxType,
        ExecutionRequest,
    )

    config = SandboxConfig(type=SandboxType.DOCKER)
    manager = SandboxManager(config)

    try:
        # Test 1: Create Docker sandbox
        print("  Test 1: Create sandbox...", end=" ")
        sandbox = await manager.get_sandbox("docker-test-1")
        assert sandbox is not None
        assert manager.get_sandbox_count() == 1
        print("✓")

        # Test 2: Execute in sandbox
        print("  Test 2: Execute in sandbox...", end=" ")
        response = await sandbox.execute(ExecutionRequest(
            code="result = 'Hello from managed sandbox'"
        ))
        assert response.success
        assert "managed sandbox" in response.output
        print("✓")

        # Test 3: Reuse sandbox
        print("  Test 3: Reuse sandbox...", end=" ")
        sandbox2 = await manager.get_sandbox("docker-test-1")
        assert sandbox is sandbox2
        assert manager.get_sandbox_count() == 1
        print("✓")

        print("  SandboxManager: All tests passed!")

    finally:
        await manager.release_all()
        assert manager.get_sandbox_count() == 0


async def main():
    """Run all Docker sandbox tests."""
    print("=" * 60)
    print("Docker Sandbox Tests")
    print("=" * 60)

    # Check arguments
    build_only = "--build-image" in sys.argv

    # Step 1: Check Docker availability
    if not await check_docker():
        print("\n❌ Docker is not available. Please install and start Docker.")
        sys.exit(1)

    # Step 2: Build image
    if not await build_image():
        print("\n❌ Failed to build Docker image.")
        sys.exit(1)

    if build_only:
        print("\n✓ Image built successfully. Use without --build-image to run tests.")
        return

    try:
        # Run tests
        await test_basic_execution()
        await test_file_io()
        await test_timeout()
        await test_package_installation()
        await test_resource_stats()
        await test_sandbox_manager()

        print("\n" + "=" * 60)
        print("ALL DOCKER SANDBOX TESTS PASSED!")
        print("=" * 60)

    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
