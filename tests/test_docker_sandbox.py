"""
Tests for Docker Sandbox.

These tests require Docker to be installed and running.
Tests are automatically skipped if Docker is not available.
"""

import asyncio
import pytest
from pathlib import Path

# Check Docker availability
try:
    import docker
    client = docker.from_env()
    client.ping()
    DOCKER_AVAILABLE = True
except Exception:
    DOCKER_AVAILABLE = False

pytestmark = pytest.mark.skipif(
    not DOCKER_AVAILABLE,
    reason="Docker not available"
)


from cape.runtime.sandbox import (
    SandboxManager,
    SandboxConfig,
    SandboxType,
    ExecutionRequest,
)

# Import DockerSandbox only if Docker is available
if DOCKER_AVAILABLE:
    from cape.runtime.sandbox import DockerSandbox


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
async def docker_sandbox():
    """Create a DockerSandbox for testing."""
    config = SandboxConfig(
        type=SandboxType.DOCKER,
        timeout_seconds=30,
        max_memory_mb=256,
    )
    sandbox = DockerSandbox(config)
    await sandbox.setup()
    yield sandbox
    await sandbox.cleanup()


@pytest.fixture
async def docker_manager():
    """Create a SandboxManager with Docker config."""
    config = SandboxConfig(type=SandboxType.DOCKER)
    manager = SandboxManager(config)
    yield manager
    await manager.release_all()


# ============================================================
# Basic Execution Tests
# ============================================================

class TestDockerBasicExecution:
    """Tests for basic code execution."""

    @pytest.mark.asyncio
    async def test_simple_calculation(self, docker_sandbox):
        """Test simple arithmetic."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="result = 2 + 2"
        ))

        assert response.success
        assert response.output == 4

    @pytest.mark.asyncio
    async def test_with_arguments(self, docker_sandbox):
        """Test execution with arguments."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="result = args['x'] * args['y']",
            args={"x": 7, "y": 8},
        ))

        assert response.success
        assert response.output == 56

    @pytest.mark.asyncio
    async def test_with_imports(self, docker_sandbox):
        """Test code with standard library imports."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
import json
data = {"name": "test", "value": 42}
result = json.dumps(data)
"""
        ))

        assert response.success
        assert "test" in response.output

    @pytest.mark.asyncio
    async def test_stdout_capture(self, docker_sandbox):
        """Test stdout capture."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
print("Hello from Docker!")
result = "done"
"""
        ))

        assert response.success
        assert "Hello from Docker!" in response.stdout

    @pytest.mark.asyncio
    async def test_error_handling(self, docker_sandbox):
        """Test error handling."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="result = 1 / 0"
        ))

        assert not response.success
        assert response.exit_code != 0


# ============================================================
# File I/O Tests
# ============================================================

class TestDockerFileIO:
    """Tests for file input/output."""

    @pytest.mark.asyncio
    async def test_file_input(self, docker_sandbox):
        """Test reading input files."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
content = Path("input.txt").read_text()
result = len(content)
""",
            files={"input.txt": b"Hello World!"},
        ))

        assert response.success
        assert response.output == 12

    @pytest.mark.asyncio
    async def test_file_output(self, docker_sandbox):
        """Test creating output files."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
Path("output.txt").write_text("Generated content")
result = "done"
"""
        ))

        assert response.success
        assert "output.txt" in response.files_created
        assert response.files_created["output.txt"] == b"Generated content"

    @pytest.mark.asyncio
    async def test_multiple_files(self, docker_sandbox):
        """Test handling multiple files."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
from pathlib import Path
import json

data1 = Path("data1.json").read_text()
data2 = Path("data2.json").read_text()
combined = {"merged": [json.loads(data1), json.loads(data2)]}
Path("result.json").write_text(json.dumps(combined))
result = len(combined["merged"])
""",
            files={
                "data1.json": b'{"a": 1}',
                "data2.json": b'{"b": 2}',
            },
        ))

        assert response.success
        assert response.output == 2
        assert "result.json" in response.files_created


# ============================================================
# Timeout Tests
# ============================================================

class TestDockerTimeout:
    """Tests for timeout handling."""

    @pytest.mark.asyncio
    async def test_timeout_kills_execution(self):
        """Test that long-running code is killed."""
        config = SandboxConfig(
            type=SandboxType.DOCKER,
            timeout_seconds=2,
        )
        sandbox = DockerSandbox(config)
        await sandbox.setup()

        try:
            response = await sandbox.execute(ExecutionRequest(
                code="""
import time
time.sleep(10)
result = "Should not reach"
"""
            ))

            assert not response.success
            assert "timeout" in response.error.lower()

        finally:
            await sandbox.cleanup()

    @pytest.mark.asyncio
    async def test_container_recovery_after_timeout(self):
        """Test container is usable after timeout."""
        config = SandboxConfig(
            type=SandboxType.DOCKER,
            timeout_seconds=2,
        )
        sandbox = DockerSandbox(config)
        await sandbox.setup()

        try:
            # Trigger timeout
            await sandbox.execute(ExecutionRequest(
                code="import time; time.sleep(10)"
            ))

            # Should still work
            response = await sandbox.execute(ExecutionRequest(
                code="result = 'recovered'"
            ))

            assert response.success
            assert response.output == "recovered"

        finally:
            await sandbox.cleanup()


# ============================================================
# Package Tests
# ============================================================

class TestDockerPackages:
    """Tests for pre-installed packages."""

    @pytest.mark.asyncio
    async def test_openpyxl(self, docker_sandbox):
        """Test openpyxl is available."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
result = ws['A1'].value
"""
        ))

        assert response.success
        assert response.output == "Hello"

    @pytest.mark.asyncio
    async def test_pandas(self, docker_sandbox):
        """Test pandas is available."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
import pandas as pd
df = pd.DataFrame({'a': [1, 2, 3], 'b': [4, 5, 6]})
result = int(df['a'].sum())
"""
        ))

        assert response.success
        assert response.output == 6

    @pytest.mark.asyncio
    async def test_numpy(self, docker_sandbox):
        """Test numpy is available."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
import numpy as np
arr = np.array([1, 2, 3, 4, 5])
result = int(arr.sum())
"""
        ))

        assert response.success
        assert response.output == 15


# ============================================================
# Manager Tests
# ============================================================

class TestDockerSandboxManager:
    """Tests for SandboxManager with Docker."""

    @pytest.mark.asyncio
    async def test_create_docker_sandbox(self, docker_manager):
        """Test creating Docker sandbox via manager."""
        sandbox = await docker_manager.get_sandbox("test-docker")

        assert sandbox is not None
        assert isinstance(sandbox, DockerSandbox)
        assert docker_manager.get_sandbox_count() == 1

    @pytest.mark.asyncio
    async def test_sandbox_reuse(self, docker_manager):
        """Test sandbox reuse."""
        sandbox1 = await docker_manager.get_sandbox("test-docker")
        sandbox2 = await docker_manager.get_sandbox("test-docker")

        assert sandbox1 is sandbox2
        assert docker_manager.get_sandbox_count() == 1

    @pytest.mark.asyncio
    async def test_execute_in_managed_sandbox(self, docker_manager):
        """Test execution in managed sandbox."""
        sandbox = await docker_manager.get_sandbox("test-docker")

        response = await sandbox.execute(ExecutionRequest(
            code="result = 'Hello from managed sandbox'"
        ))

        assert response.success
        assert "managed sandbox" in response.output


# ============================================================
# Integration Tests
# ============================================================

class TestDockerIntegration:
    """Integration tests for Docker sandbox."""

    @pytest.mark.asyncio
    async def test_excel_creation(self, docker_sandbox):
        """Test creating an Excel file."""
        response = await docker_sandbox.execute(ExecutionRequest(
            code="""
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Value'
ws['A2'] = 'Test'
ws['B2'] = 42
ws['B3'] = '=SUM(B2:B2)'

wb.save('output.xlsx')
result = "Excel file created"
"""
        ))

        assert response.success
        assert "output.xlsx" in response.files_created
        assert len(response.files_created["output.xlsx"]) > 0

    @pytest.mark.asyncio
    async def test_data_processing_workflow(self, docker_sandbox):
        """Test a data processing workflow."""
        # Step 1: Create data
        response1 = await docker_sandbox.execute(ExecutionRequest(
            code="""
import json
from pathlib import Path

data = [
    {"name": "Alice", "score": 85},
    {"name": "Bob", "score": 92},
    {"name": "Charlie", "score": 78},
]

Path("data.json").write_text(json.dumps(data))
result = len(data)
"""
        ))

        assert response1.success
        assert response1.output == 3

        # Step 2: Process data in new execution (files don't persist)
        response2 = await docker_sandbox.execute(ExecutionRequest(
            code="""
import json
from pathlib import Path

# Read input
data = json.loads(Path("data.json").read_text())

# Process
avg_score = sum(d["score"] for d in data) / len(data)
result = {"average_score": avg_score, "count": len(data)}

# Save results
Path("results.json").write_text(json.dumps(result))
""",
            files={"data.json": response1.files_created["data.json"]},
        ))

        assert response2.success
        assert response2.output["count"] == 3
        assert response2.output["average_score"] == 85.0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
